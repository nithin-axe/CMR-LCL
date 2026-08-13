"""Identify the trade-document type of an email and its attachments.

Two classification paths exist, by design:

* ``classify_email_meta`` - READ-SAFE. Uses only list-view metadata (sender, subject,
  snippet, attachment file names). Never opens/downloads a mail, so it can't change the
  message's read/unread state. This backs the bulk "Find Document Types" button, which
  must be safe to run against the delegated (scraped) mailbox.

* ``classify_email`` / ``classify_documents`` - DEEP. Reads the actual document bytes
  (PDF/image) and the full body, sends them to Gemini multimodal, and can therefore:
    1. read a document-type keyword printed INSIDE the document (preferred signal), and
    2. detect that a single PDF spans MULTIPLE sub-documents across its pages.
  Opening/downloading a delegated-mailbox mail can mark it read, so this path is only
  used from the explicit "Deep analyze" action, never the bulk sweep.

Classification priority (both paths, encoded in the prompts):
  1. An explicit document-type keyword found IN the document (deep) or in the attachment
     file name / subject (meta).
  2. Failing that, infer from the email subject + body/preview text.

Canonical output types and the name->type reference map below are the single source of
truth. Results are cached in ``data/doc_classifications.json`` keyed by message id.
"""

import base64
import concurrent.futures
import json
import os
import re
import threading
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from flask import current_app

from app.services.common.email_service import EmailService
from app.services.llm.llm_client import GeminiClient


# The lcl-arrivals---release label (see tracking_api.py's _LCL_LABEL_KEY) passed as
# classify_email_meta/resolve_deep_classification's ``label`` to select the label
# override path below - same mechanism as "a-cmr", but this one restricts the output
# to ONLY "Arrival notice"/"Delivery order"/"Other", never the full CMR-oriented
# DOCUMENT_TYPES list. Duplicated as a literal here rather than imported, matching how
# "a-cmr" is already hardcoded inline rather than shared from tracking_api.py.
_LCL_ARRIVALS_LABEL = "lcl-arrivals---release"

# classify_lcl_arrival_email's 4-way mail-type result, mapped onto the 2 document
# types this label's "Find Document Types" is allowed to show - shipment_not_released/
# delay_or_devanning/unknown mails (and anything else) fall back to "Other", exactly
# like a-cmr's non-Cmr attachments do.
_LCL_MAIL_TYPE_TO_DOC_TYPE = {
    "arrival_notice": "Arrival notice",
    "delivery_order": "Delivery order",
}

# Canonical document types - the ONLY values a classification may take. Taken from the
# operator's reference table (the right-hand "Document type" column), plus the four
# extra types called out in the row-by-row corrections (T1 DOC, OR FILES, TRANSFER,
# EUR AND PHYTO). The LLM is constrained to copy exactly one of these verbatim.
DOCUMENT_TYPES = [
    "Final master bill of lading",
    "Commercial invoice",
    "Phytosanitary certificate",
    "EUR1 certificate",
    "Delivery order",
    "MRN",
    "Chedpp",
    "Inspection report",
    "Arrival notice",
    "Cmr",
    "Invoice Declration",
    "Custome Import Doc",
    "custome duties And Vat",
    "Cargo Manifest",
    "Certificate Of Inspection (COI)",
    "Packing List",
    "Certificate Of Orgin",
    "No DOC",
    "Release House BL",
    "Draft HBL",
    "final Value Declration",
    "Release Master Bill",
    "Draft MBL",
    "T1 DOC",
    "OR FILES",
    "TRANSFER",
    "EUR AND PHYTO",
    "Other",
]

# Types whose real content typically lists container numbers. The metadata-only live
# pass (classify_email_meta) can only see numbers that happen to leak into the
# subject/snippet/attachment filename - for these types, if that pass found none, the
# automatic live pipeline escalates to a real content read (see
# tracking_api.classify_new_documents_cmr/_lcl) rather than leaving the row's container chip
# silently empty just because the number is buried inside the attachment. Deliberately
# narrow (invoices/certificates/customs paperwork are excluded) so the automatic
# escalation - which opens the mail and marks it read - only fires for types that are
# actually likely to be carrying containers.
CONTAINER_BEARING_TYPES = {
    "Final master bill of lading",
    "Release Master Bill",
    "Draft MBL",
    "Release House BL",
    "Draft HBL",
    "Delivery order",
    "Arrival notice",
    "Cmr",
    "Cargo Manifest",
    "T1 DOC",
    "Packing List",
}

# The operator's reference table: how a recognised document NAME maps to the canonical
# output type. Injected verbatim into every prompt so the LLM applies these exact rules
# instead of guessing. (document name / description  ->  canonical type)
REFERENCE_MAP = [
    ("Final master bill of lading / original sheet", "Final master bill of lading"),
    ("Commercial invoice", "Commercial invoice"),
    ("Phytosanitary (Phytosanitairy) certificate", "Phytosanitary certificate"),
    ("EUR1", "EUR1 certificate"),
    ("Delivery order carrier / RELEASE ORDER", "Delivery order"),
    ("NCTS release", "MRN"),
    ("Chedpp", "Chedpp"),
    ("Inspectierapport (inspection report)", "Inspection report"),
    ("Nieuwe berichten Naktuinbouw", "Inspection report"),
    ("Arrival notice", "Arrival notice"),
    ("Cmr", "Cmr"),
    ("Commercial invoice agricola", "Invoice Declration"),
    ("M1", "Custome Import Doc"),
    ("Amount Sheet", "custome duties And Vat"),
    ("Cargo Manifest", "Cargo Manifest"),
    ("C.O.I", "Certificate Of Inspection (COI)"),
    ("PL / Packing List", "Packing List"),
    ("Export Certificate / Certificate Of Origin", "Certificate Of Orgin"),
    ("DHL", "Other"),
    ("Final account", "Commercial invoice"),
    ("Pin Sheet", "Delivery order"),
    ("MEDEDELING STATUS VOORAANMELDING", "Other"),
    ("HBL", "Release House BL"),
    ("Draft HBL", "Draft HBL"),
    ("Balance Payment", "final Value Declration"),
    ("Factura", "Commercial invoice"),
    ("Non Way Bill / Way Bill / Sea Way Bill / release Sheet", "Release Master Bill"),
    ("Draft MBL", "Draft MBL"),
    ("Split List", "Other"),
    # Row-by-row corrections (Deepak.txt) that introduce / override with the new types:
    ('"A release has been transferred" notification (securecontainerrelease.com)', "TRANSFER"),
    ("Automatic write-off of NCTS5_DEPARTURE / T1 transit document", "T1 DOC"),
    ("Automatic mail for the release of IDMS_IMPORT / Customs Import Document", "Custome Import Doc"),
    ("A bundle of release / order (O.R.) files", "OR FILES"),
    ("A mail carrying BOTH an EUR1 and a Phytosanitary certificate", "EUR AND PHYTO"),
]

# Concrete sender/subject examples from the operator's corrections. Given to the LLM as
# few-shot guidance so recurring notification mails land on the right type even from
# metadata alone.
_FEW_SHOT_EXAMPLES = [
    ('DVK Transport - "Container: TEMU 968174-4, Plaats: RIJNSBURG"', "Cmr"),
    ('NETHERLANDS - "IMCR0101 - M;HLCULI3260631361;MD;ECT ;001"', "Delivery order"),
    ('noreply - "A release has been transferred for BL MEDUJ0755564"', "TRANSFER"),
    ('stella.keymolen - "Automatic mail for the write off of NCTS5_DEPARTURE ..."', "T1 DOC"),
    ('luisa.meziat-pina-v - "Automatic mail for the release of IDMS_IMPORT with 26BEH..."', "Custome Import Doc"),
    ('KCB BOSS - "Rapportage uitgevoerde inspectie: NL8135... "', "Inspection report"),
    ('Plantkeur/KCB - "KCB: NL8135... status vooraanmelding CXRU1450965"', "Other"),
    ('no-reply - "ISB2008962 ... NOA-CMACGM-... Notice of Arrival"', "OR FILES"),
]

# Fast filename keyword hints. Ordered most-specific first so e.g. "draft mbl" wins over
# "mbl". Each entry: (regex over the lowercased filename) -> canonical type. A hint is
# only ever a suggestion passed to the LLM / a fallback, never the final answer alone.
_FILENAME_HINTS = [
    (r"draft.*mbl", "Draft MBL"),
    (r"draft.*hbl", "Draft HBL"),
    (r"\bhbl\b|house.*b/?l", "Release House BL"),
    (r"\bmbl\b|master.*b/?l|bill.*of.*lading|\bb/?l\b|original\s*sheet", "Final master bill of lading"),
    (r"phyto", "Phytosanitary certificate"),
    (r"eur[\s._-]?1", "EUR1 certificate"),
    # NCTS / IDMS covers transit & import release notifications that map to specific canonical types:
    (r"idms.*import|import.*idms", "Custome Import Doc"),
    (r"\bt1\b|ncts.*(write.?off|departure)|(write.?off|departure).*ncts", "T1 DOC"),
    (r"ncts.*release|release.*ncts|\bmrn\d*\b", "MRN"),
    (r"chedpp|chep", "Chedpp"),
    (r"inspection|inspectie|rapportage|naktuinbouw", "Inspection report"),
    (r"arrival\s*notice|\bnoa\b|notice\s*of\s*arrival", "Arrival notice"),
    (r"\bcmr\b", "Cmr"),
    (r"agricola|agricole", "Invoice Declration"),
    (r"\bm1\b", "Custome Import Doc"),
    (r"amount\s*sheet|duties|\bvat\b", "custome duties And Vat"),
    (r"\bmanifest\b", "Cargo Manifest"),
    (r"c[\s._-]?o[\s._-]?i\b|certificate.*inspection", "Certificate Of Inspection (COI)"),
    (r"packing|\bpl\b|pack.*list", "Packing List"),
    (r"export.*cert|certificate.*orig|\bcoo\b|\bc/?o\b", "Certificate Of Orgin"),
    (r"final.*account", "Commercial invoice"),
    (r"pin\s*sheet", "Delivery order"),
    (r"mededeling|vooraanmelding", "Other"),
    (r"release.*transfer|transfer.*release", "TRANSFER"),
    (r"delivery\s*order|release\s*order|\bdo\b", "Delivery order"),
    (r"balance.*payment", "final Value Declration"),
    (r"factura", "Commercial invoice"),
    (r"way\s*bill|waybill|swb|seaway|release\s*sheet", "Release Master Bill"),
    (r"split\s*list", "Other"),
    (r"\bor[\s._-]?files?\b", "OR FILES"),
    (r"commercial.*invoice|\binvoice\b|\binv\b", "Commercial invoice"),
]

# MIME types Gemini can read inline as documents/images. Anything else (docx, xlsx,
# zip...) is classified from its filename + surrounding email text only, UNLESS it's one
# of the Excel types below, which get their cell content extracted as text instead (see
# _extract_excel_text) - Gemini's multimodal API has no inline part type for spreadsheets.
_LLM_READABLE_PREFIXES = ("application/pdf", "image/", "text/")

_EXCEL_MIME_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel.sheet.macroEnabled.12",  # .xlsm
    "application/vnd.ms-excel",  # legacy .xls
)


def _is_excel_attachment(filename, mime):
    """Mime alone isn't always trustworthy - the delegated-mailbox scrape path
    (open_gmail.py) sets mime from the HTTP response's content-type header, which some
    servers report as a generic application/octet-stream for attachments. Fall back to
    the filename extension so Excel detection works from either source."""
    if mime in _EXCEL_MIME_TYPES:
        return True
    return (filename or "").lower().endswith((".xlsx", ".xlsm", ".xls"))

_CACHE_LOCK = threading.Lock()


def _cache_path():
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "data"))
    return os.path.join(base, "doc_classifications.json")


def _load_cache():
    path = _cache_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_cache(cache):
    path = _cache_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def clear_classification_cache():
    """Clear all saved document classifications from disk and memory, plus every
    cached attachment file (see _ATTACHMENT_CACHE_DIR below) - otherwise a "Clear
    Cache" that only wipes the type/container metadata would still silently serve
    stale bytes for a message that gets re-classified afterward."""
    path = _cache_path()
    with _CACHE_LOCK:
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
        if os.path.isdir(_ATTACHMENT_CACHE_DIR):
            try:
                import shutil
                shutil.rmtree(_ATTACHMENT_CACHE_DIR)
            except Exception:
                pass


def _delete_cached_attachments(message_id):
    """Remove every cached attachment .bin/.meta.json for one message (see
    _cache_attachment_bytes) - a message can have several, one per attachment index,
    so this globs by the id prefix rather than needing to know the count. Best-effort,
    same as clear_classification_cache's own attachment cleanup."""
    if not os.path.isdir(_ATTACHMENT_CACHE_DIR):
        return
    safe_id = re.sub(r"[^A-Za-z0-9_.-]", "_", message_id or "")
    if not safe_id:
        return
    prefix = safe_id + "_"
    try:
        for name in os.listdir(_ATTACHMENT_CACHE_DIR):
            if name.startswith(prefix):
                try:
                    os.remove(os.path.join(_ATTACHMENT_CACHE_DIR, name))
                except Exception:
                    pass
    except Exception:
        pass


def delete_classification(message_id):
    """Forget one message's cached classification (+ its cached attachment bytes, if
    any). Used by the History view's per-row delete - lets the operator discard a
    wrong/stale classification (or force a genuinely fresh one on the next Find
    Document Types / Process, instead of paying force=True's re-open-and-re-classify
    cost on a mail whose cached result was fine). Returns True if an entry actually
    existed to delete."""
    with _CACHE_LOCK:
        cache = _load_cache()
        existed = cache.pop(message_id, None) is not None
        if existed:
            _save_cache(cache)
    _delete_cached_attachments(message_id)
    return existed


def clear_classifications(message_ids):
    """Bulk delete_classification over a whole set of message ids (+ their cached
    attachment bytes) - used by the History view's per-label "Clear all". Returns how
    many actually had a cached entry to remove."""
    ids = set(message_ids or [])
    removed = 0
    with _CACHE_LOCK:
        cache = _load_cache()
        for mid in list(cache.keys()):
            if mid in ids:
                del cache[mid]
                removed += 1
        if removed:
            _save_cache(cache)
    for mid in ids:
        _delete_cached_attachments(mid)
    return removed


# Local disk cache of raw attachment BYTES, keyed by (message_id, attachment_index) -
# separate from doc_classifications.json above, which only ever stores classification
# metadata (type/containers/etc.), never the bytes themselves. Every deep classification
# already has to open the mail and download each attachment once (to feed Gemini) - this
# persists that same download to disk so the LATER Shypple upload step (fetch_document_
# bytes below) can read it straight off disk instead of paying for a second full
# open-the-mail-again round-trip through the delegated-mailbox browser, which is exactly
# what was making the upload step slow (each document taking ~12-20s AFTER the Shypple
# browser was already open, per the operator's own logs). Best-effort throughout: a
# caching failure must never break classification or the upload itself, only lose the
# speed-up for that one document.
_ATTACHMENT_CACHE_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "attachment_cache")
)


def _attachment_cache_paths(message_id, attachment_index):
    safe_id = re.sub(r"[^A-Za-z0-9_.-]", "_", message_id or "")
    base = os.path.join(_ATTACHMENT_CACHE_DIR, f"{safe_id}_{attachment_index}")
    return base + ".bin", base + ".meta.json"


def _cache_attachment_bytes(message_id, attachment_index, data_bytes, filename="", mime=""):
    if data_bytes is None or attachment_index is None or not message_id:
        return
    try:
        os.makedirs(_ATTACHMENT_CACHE_DIR, exist_ok=True)
        bin_path, meta_path = _attachment_cache_paths(message_id, attachment_index)
        with open(bin_path, "wb") as f:
            f.write(data_bytes)
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump({"filename": filename or "", "mime": mime or ""}, f)
    except Exception as e:
        _log_warn(f"Could not cache attachment bytes for {message_id}#{attachment_index}: {e}")


def _load_cached_attachment(message_id, attachment_index):
    """Returns (data_bytes, mime, filename), or (None, None, None) if nothing is cached
    for this (message_id, attachment_index) yet."""
    if attachment_index is None or not message_id:
        return None, None, None
    bin_path, meta_path = _attachment_cache_paths(message_id, attachment_index)
    if not os.path.exists(bin_path):
        return None, None, None
    try:
        with open(bin_path, "rb") as f:
            data_bytes = f.read()
    except Exception as e:
        _log_warn(f"Could not read cached attachment for {message_id}#{attachment_index}: {e}")
        return None, None, None
    filename, mime = "", ""
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            filename = meta.get("filename", "")
            mime = meta.get("mime", "")
        except Exception:
            pass
    return data_bytes, mime, filename


def _log_warn(msg):
    try:
        if current_app:
            current_app.logger.warning(msg)
            return
    except Exception:
        pass
    print(msg)


def _html_to_text(html):
    """Reduce an HTML email body to readable plain text for classification."""
    if not html:
        return ""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style"]):
            tag.decompose()
        text = soup.get_text(separator=" ")
    except Exception:
        text = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", text).strip()


def _guess_mime(filename):
    import mimetypes
    return mimetypes.guess_type(filename or "")[0] or "application/octet-stream"


def filename_hint(filename):
    """Return the best canonical type guessed purely from a filename/subject, or None."""
    # Underscores are regex word chars, so "MBL_123" would defeat a \bmbl\b boundary;
    # treat them as separators.
    name = (filename or "").lower().replace("_", " ")
    for pattern, doc_type in _FILENAME_HINTS:
        if re.search(pattern, name):
            return doc_type
    return None


# ISO 6346 container number: 3-letter owner code + category (U/J/Z) + 6 serial digits +
# 1 check digit, e.g. TEMU9681744. Real mails write them loosely ("TEMU 968174-4",
# "MRKU 952560/5", "CXRU1450965"), so allow optional space/dash/slash separators.
_CONTAINER_RE = re.compile(r"\b([A-Z]{3}[UJZ])[\s\-/]?(\d{6})[\s\-/]?(\d)\b")
# Strict ISO-6346 shape used to validate anything (esp. loose LLM output) before we keep
# it - filters out BL/booking/KCB refs like "ANT2009416" or "NL813564888.002".
_CONTAINER_VALID = re.compile(r"^[A-Z]{3}[UJZ]\d{7}$")


def find_container_numbers(text):
    """Extract normalised ISO-6346 container numbers (LETTERS+7 digits) from text."""
    if not text:
        return []
    found = []
    for m in _CONTAINER_RE.finditer(text.upper()):
        num = m.group(1) + m.group(2) + m.group(3)
        if num not in found:
            found.append(num)
    return found


def _merge_containers(*lists):
    """Merge + de-dupe container numbers from several sources, keeping only values that
    match the strict ISO-6346 shape (guards against the LLM returning BL/booking refs)."""
    out = []
    for lst in lists:
        for c in lst or []:
            c = re.sub(r"[\s\-/.]", "", (c or "").strip().upper())
            if _CONTAINER_VALID.match(c) and c not in out:
                out.append(c)
    return out


def _extract_pdf_page_texts(data_bytes):
    """Extract each page's text layer from a PDF, for a DETERMINISTIC, regex-based
    container-number pass that doesn't depend on the LLM noticing every entry in a long
    table. Returns ``[]`` for anything that isn't a text-based PDF (encrypted, corrupt, or
    a scanned image with no text layer) - those still get container numbers from the
    LLM's multimodal reading, this is a supplementary safety net, not a replacement."""
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(data_bytes))
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                return []
        return [(page.extract_text() or "") for page in reader.pages]
    except Exception as e:
        _log_warn(f"PDF text extraction failed (falling back to LLM-only container reading): {e}")
        return []


def _extract_excel_text(data_bytes, max_rows_per_sheet=200, max_sheets=8, max_chars=12000):
    """Render a workbook's actual cell content as plain text so Gemini - which can't read
    .xlsx/.xlsm/.xls inline the way it reads a PDF or image - can still classify it from
    its real content rather than just the filename. Capped hard on rows/sheets/total
    length: a real workbook (e.g. an activity log) can run to tens of thousands of rows,
    and this is a classification aid fed into a prompt, not a data export."""
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data_bytes), data_only=True, read_only=True)
    except Exception as e:
        _log_warn(f"Excel text extraction failed (unreadable/corrupt/password-protected): {e}")
        return ""

    lines = []
    try:
        for sheet_name in wb.sheetnames[:max_sheets]:
            ws = wb[sheet_name]
            sheet_lines = [f"--- Sheet: {sheet_name} ---"]
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count >= max_rows_per_sheet:
                    sheet_lines.append("... (truncated, more rows follow)")
                    break
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if not cells:
                    continue
                sheet_lines.append(" | ".join(cells))
                row_count += 1
            if row_count:
                lines.extend(sheet_lines)
            if sum(len(l) for l in lines) > max_chars:
                break
    except Exception as e:
        _log_warn(f"Excel sheet iteration failed partway through: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return "\n".join(lines)[:max_chars]


def _parse_page_range(pages_str, total_pages):
    """Parse a page-range string like '1-2', '3', '1,3-4', 'all' into a 0-indexed list of
    page numbers into ``page_texts``. Unparseable/'all' -> every page (safest default)."""
    if not pages_str or not total_pages:
        return list(range(total_pages))
    s = str(pages_str).strip().lower()
    if s in ("all", "-", ""):
        return list(range(total_pages))
    pages = set()
    try:
        for part in s.split(","):
            part = part.strip()
            if "-" in part:
                a, b = part.split("-", 1)
                a, b = int(a), int(b)
                for p in range(min(a, b), max(a, b) + 1):
                    if 1 <= p <= total_pages:
                        pages.add(p - 1)
            elif part.isdigit():
                p = int(part)
                if 1 <= p <= total_pages:
                    pages.add(p - 1)
    except (ValueError, TypeError):
        return list(range(total_pages))
    return sorted(pages) if pages else list(range(total_pages))


def containers_from_pdf_pages(page_texts, pages_str):
    """Deterministic container numbers for the page range a sub-document covers, scanned
    directly from the PDF's own text layer (not the LLM's summary of it)."""
    if not page_texts:
        return []
    indices = _parse_page_range(pages_str, len(page_texts))
    text = " ".join(page_texts[i] for i in indices if 0 <= i < len(page_texts))
    return find_container_numbers(text)


# Shared guardrail text - stops the model inventing a specific type (e.g. Arrival notice)
# for a mail that carries no actual document or only generic shipping references.
_GUARDRAIL = (
    "AVOID FALSE POSITIVES:\n"
    "- Assign a specific document type ONLY when the mail clearly IS or CONTAINS that document.\n"
    "- A container number, booking/BL reference, ETA or generic shipping wording is NOT by\n"
    "  itself an 'Arrival notice', 'Delivery order', etc.\n"
    "- If the mail has no attached document and the text does not clearly present one of the\n"
    "  listed document types, answer \"No DOC\".\n"
    "- Never guess 'Arrival notice' just because shipping details / a container are mentioned.\n"
    "- 'Cargo Manifest' means the document IS an actual cargo/shipment manifest listing (a\n"
    "  table of shipment line items). A filename that merely CONTAINS the word \"manifest\", or\n"
    "  a passing mention of a manifest reference number in the text, is NOT enough on its own -\n"
    "- MRN vs T1 DOC vs Custome Import Doc:\n"
    "  * An email titled \"Automatic mail for the release of IDMS_IMPORT ...\" -> \"Custome Import Doc\".\n"
    "  * An NCTS WRITE-OFF / DEPARTURE notification (\"Automatic mail for the write off of NCTS5_DEPARTURE ...\") -> \"T1 DOC\".\n"
    "  * An NCTS RELEASE notification (where no printed declaration form or IDMS_IMPORT is attached) -> \"MRN\".\n"
    "- EU customs declaration FORM / Accompanying Document (the official printed form with an MRN barcode plus box [1101]):\n"
    "  * \"T1\" in box [1101] -> \"T1 DOC\".\n"
    "  * \"IM\" in box [1101] -> \"Custome Import Doc\".\n"
    "  * Do NOT default to \"MRN\" just because of the MRN barcode header - the MRN number appears on both T1 and IM declarations.\n"
)


def _numbered_types():
    return "\n".join(f"{i + 1}. {t}" for i, t in enumerate(DOCUMENT_TYPES))


def _reference_block():
    lines = ["Reference mapping (recognised document name -> the type you must output):"]
    for name, canon in REFERENCE_MAP:
        lines.append(f"  - {name}  ->  {canon}")
    return "\n".join(lines)


def _few_shot_block():
    lines = ["Worked examples (email -> correct type):"]
    for example, canon in _FEW_SHOT_EXAMPLES:
        lines.append(f"  - {example}  ->  {canon}")
    return "\n".join(lines)


def _parse_llm_json(text):
    if not text:
        return None
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?|```$", "", cleaned, flags=re.MULTILINE).strip()
    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    if not match:
        return None
    try:
        return json.loads(match.group(0))
    except Exception:
        return None


def _coerce_type(value):
    """Snap an LLM-returned type string onto the canonical list (exact, then loose)."""
    if not value:
        return None
    v = value.strip()
    for t in DOCUMENT_TYPES:
        if t.lower() == v.lower():
            return t
    norm = re.sub(r"[^a-z0-9]", "", v.lower())
    for t in DOCUMENT_TYPES:
        if re.sub(r"[^a-z0-9]", "", t.lower()) == norm:
            return t
    return None


# --------------------------------------------------------------------------- #
# DEEP path - reads the real document bytes (keyword-in-document first, then the
# surrounding email text), and supports one PDF holding several sub-documents.
# --------------------------------------------------------------------------- #

def _build_body_prompt(subject, body_text):
    return (
        "You classify logistics / import trade documents for a freight-forwarding workflow.\n"
        "Choose the SINGLE best matching type for the EMAIL BODY itself from this list:\n\n"
        f"{_numbered_types()}\n\n"
        f"{_reference_block()}\n\n"
        f"{_few_shot_block()}\n\n"
        "Rules:\n"
        "- The \"type\" MUST be copied verbatim from the numbered list (exact spelling).\n"
        "- Decide from the body text meaning.\n\n"
        f"{_GUARDRAIL}\n"
        "Also extract EVERY ISO-6346 container number mentioned (4 letters + 7 digits, e.g.\n"
        "TEMU9681744) into \"containers\". The body may list several containers (e.g. a\n"
        "table or bullet list) - scan the WHOLE text and include ALL of them, not just the\n"
        "first one or two.\n\n"
        f"Email subject: {subject or '(none)'}\n"
        f"Email body text:\n{(body_text or '')[:6000]}\n\n"
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"type": "<one exact value>", "confidence": <0.0-1.0>, "reason": "<short>", "containers": ["..."]}'
    )


def _build_attachment_prompt(subject, body_text, filename, hint, excel_text=""):
    hint_line = (
        f"A crude filename keyword scan guessed \"{hint}\" - this looks ONLY at the file name,\n"
        f"not the content, and is frequently WRONG (e.g. it cannot tell a Cargo Manifest from a\n"
        f"file that merely has \"manifest\" in its name, or an MRN release from a T1 write-off/\n"
        f"departure notice). Do NOT default to this guess - agree with it only if the document's\n"
        f"actual content independently confirms it; otherwise ignore it and decide from content.\n"
        if hint else ""
    )
    excel_block = (
        "This attachment is a SPREADSHEET. Gemini cannot read .xlsx/.xlsm/.xls file bytes\n"
        "inline, so its cell contents have been extracted as plain text below - read it the\n"
        "same way you would read the pages of a PDF, and classify/extract containers from it:\n"
        f"{excel_text}\n\n"
        if excel_text else ""
    )
    return (
        "You classify logistics / import trade documents for a freight-forwarding workflow.\n"
        "You are given the ACTUAL document (PDF/image, or extracted text for a spreadsheet)\n"
        "plus its surrounding email text.\n\n"
        f"Allowed types (copy verbatim, exact spelling):\n{_numbered_types()}\n\n"
        f"{_reference_block()}\n\n"
        f"{_few_shot_block()}\n\n"
        "HOW TO DECIDE (in this order):\n"
        "1. KEYWORD IN THE DOCUMENT: read the document and look for an explicit type/title\n"
        "   printed on it (e.g. 'CMR', 'PHYTOSANITARY CERTIFICATE', 'BILL OF LADING',\n"
        "   'EUR.1', 'PACKING LIST', 'DELIVERY ORDER'). If the document states its own\n"
        "   type, map it through the reference table and use that.\n"
        "2. ONLY IF the document has no such keyword, infer the type from the email\n"
        "   subject and body text.\n\n"
        "MULTIPLE PAGES / MULTIPLE DOCUMENTS: a single PDF very often contains several\n"
        "different documents across its pages (e.g. pages 1-2 an Inspection report, page 3 a\n"
        "Sampling registration report, page 4 an analysis certificate). You MUST read EVERY\n"
        "page to the end and return ONE entry per DISTINCT document type you find, each with\n"
        "its page range. Do not stop after the first document. Only return a single entry if\n"
        "the whole file genuinely is one document.\n\n"
        f"{_GUARDRAIL}\n"
        "CONTAINER NUMBERS: for EACH document entry, extract EVERY ISO-6346 container\n"
        "number printed on those pages (4 letters + 7 digits, e.g. TEMU9681744) into its\n"
        "\"containers\" array. A single document can list MANY containers (a manifest, a\n"
        "release order, or a table can have dozens) - read the ENTIRE page range carefully\n"
        "and list every one you can find, in the order they appear. Do not stop after the\n"
        "first few or summarize - completeness matters more than brevity here.\n\n"
        f"Email subject: {subject or '(none)'}\n"
        f"Attachment file name: {filename or '(unnamed)'}\n"
        f"{hint_line}"
        f"{excel_block}"
        f"Surrounding email body (context only):\n{(body_text or '')[:1500]}\n\n"
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"documents": [{"type": "<one exact value>", "pages": "<e.g. 1-2 or all>", '
        '"confidence": <0.0-1.0>, "reason": "<short, mention the keyword you saw>", "containers": ["..."]}]}'
    )


def _classify_body(gemini, subject, body_text):
    prompt = _build_body_prompt(subject, body_text)
    doc_type, confidence, reason, method, containers = None, 0.0, "", "llm", []
    try:
        parsed = _parse_llm_json(gemini.generate_multimodal(prompt, []))
        if parsed:
            doc_type = _coerce_type(parsed.get("type"))
            confidence = _safe_float(parsed.get("confidence"))
            reason = str(parsed.get("reason", ""))[:300]
            containers = parsed.get("containers") or []
    except Exception as e:
        _log_warn(f"Body classification LLM call failed: {e}")

    if not doc_type:
        hint = filename_hint(subject)
        if hint:
            doc_type, method, confidence = hint, "keyword", 0.4
            reason = reason or "Matched by subject keyword."
        else:
            doc_type, method, confidence = "No DOC", "fallback", 0.0
            reason = reason or "No document present."
    containers = _merge_containers(containers, find_container_numbers(f"{subject} {body_text}"))
    return {"source": "body", "type": doc_type, "confidence": round(confidence, 2),
            "method": method, "reason": reason, "pages": "-", "containers": containers}


def _classify_attachment(gemini, subject, body_text, filename, data_bytes, mime):
    """Classify one attachment. Returns a LIST of result dicts - one per distinct
    document type found across the file's pages.

    Container numbers combine TWO independent sources so nothing gets missed:
      1. The LLM's own reading (works on scanned/image-only pages too), and
      2. A deterministic regex pass over the PDF's actual text layer, scoped to each
         sub-document's page range - this catches every entry even in a long table where
         the model's attention might only report the first few."""
    hint = filename_hint(filename)

    # Excel can't be read inline by Gemini - extract its cell content as text instead and
    # feed that into the prompt as real content, not just a filename guess.
    excel_text = ""
    if data_bytes and _is_excel_attachment(filename, mime):
        excel_text = _extract_excel_text(data_bytes)

    prompt = _build_attachment_prompt(subject, body_text, filename, hint, excel_text)

    parts = []
    readable = bool(data_bytes) and mime.startswith(_LLM_READABLE_PREFIXES)
    if readable and len(data_bytes) <= 18 * 1024 * 1024:  # Gemini inline cap ~20MB
        parts.append((mime, data_bytes))

    # Deterministic safety net: extract every page's real text layer once, up front.
    page_texts = []
    if data_bytes and mime == "application/pdf":
        page_texts = _extract_pdf_page_texts(data_bytes)
    whole_doc_containers = find_container_numbers(" ".join(page_texts)) if page_texts else []
    if excel_text:
        whole_doc_containers = _merge_containers(whole_doc_containers, find_container_numbers(excel_text))

    subdocs = []
    try:
        parsed = _parse_llm_json(gemini.generate_multimodal(prompt, parts))
        if parsed and isinstance(parsed.get("documents"), list):
            for d in parsed["documents"]:
                t = _coerce_type(d.get("type"))
                if not t:
                    continue
                pages = str(d.get("pages", "all"))[:20]
                page_scoped = containers_from_pdf_pages(page_texts, pages) if page_texts else []
                extra_containers = find_container_numbers(excel_text) if excel_text else []
                subdocs.append({
                    "source": "attachment",
                    "filename": filename,
                    "type": t,
                    "pages": pages,
                    "confidence": round(_safe_float(d.get("confidence")), 2),
                    "method": "content" if readable else ("excel-text" if excel_text else "llm"),
                    "reason": str(d.get("reason", ""))[:300],
                    "containers": _merge_containers(d.get("containers"), page_scoped, extra_containers),
                })
    except Exception as e:
        _log_warn(f"Attachment classification LLM call failed for '{filename}': {e}")

    if not subdocs:
        base_containers = _merge_containers(find_container_numbers(filename), whole_doc_containers)
        if hint:
            subdocs.append({"source": "attachment", "filename": filename, "type": hint,
                            "pages": "all", "confidence": 0.4, "method": "keyword",
                            "reason": "Matched by filename keyword.",
                            "containers": base_containers})
        else:
            subdocs.append({"source": "attachment", "filename": filename, "type": "Other",
                            "pages": "all", "confidence": 0.0, "method": "fallback",
                            "reason": "No confident match.",
                            "containers": base_containers})
    return subdocs


def _safe_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _distinct_types(types):
    """Dedupe a list of types (order-preserving). Keep 'Other' alongside specific
    types like 'Cmr' so non-CMR attachments are extracted and tracked."""
    seen, ordered = set(), []
    for t in types:
        if t and t not in seen:
            seen.add(t)
            ordered.append(t)
    meaningful = [t for t in ordered if t != "No DOC"]
    return meaningful if meaningful else ordered[:1]


def _type_counts(attachment_types):
    """Tally how many separate document instances share each type, counting only
    attachment-sourced entries (not the email body). Includes 'Other'."""
    counts = {}
    for t in attachment_types:
        if t and t != "No DOC":
            counts[t] = counts.get(t, 0) + 1
    return counts


def classify_documents(message_id, subject, body_html, attachments, force=False):
    """Core deep classifier shared by the Gmail-API path and the delegated-mailbox
    deep-scan. ``attachments`` is a list of dicts with keys: ``filename``, ``mime``
    (optional - guessed from the name if absent) and ``data_bytes`` (may be None).

    Returns ``{"message_id", "email_type", "documents": [...], "source": "deep"}``.
    """
    cache = _load_cache()
    if not force and message_id in cache and cache[message_id].get("source") == "deep":
        cached = _backfill_missing_cmr_containers(message_id, dict(cache[message_id]))
        cached["cached"] = True
        return cached

    body_text = _html_to_text(body_html or "")
    gemini = GeminiClient()

    # The body classification and each attachment's classification are independent
    # Gemini calls - running them one after another (as this used to) means an
    # N-attachment mail pays N+1 sequential LLM round-trips (each several seconds),
    # which is where most of a "deep" classification's wall-clock time goes for any
    # label without a-cmr's type-override shortcut (classify_documents_cmr skips the
    # per-attachment LLM call entirely for the overridden Cmr/Other slot - this
    # generic path, used by lcl-arrivals---release and any other label, has no such
    # shortcut and was measurably the slower of the two as a result). Run them
    # concurrently instead - same calls, same results, just not serialized.
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(attachments or []) + 1) as pool:
        body_future = pool.submit(_classify_body, gemini, subject, body_text)
        att_futures = [
            (i, pool.submit(
                _classify_attachment, gemini, subject, body_text,
                att.get("filename", "") or "", att.get("data_bytes"),
                att.get("mime") or _guess_mime(att.get("filename", "") or ""),
            ))
            for i, att in enumerate(attachments or [])
        ]

        documents = [body_future.result()]
        for i, fut in att_futures:
            sub_docs = fut.result()
            # Recorded so fetch_document_bytes can re-fetch this exact attachment by
            # POSITION later, instead of by filename text - the filename extracted at
            # classify time can be a generic fallback ("Attachment 1", used when the real
            # name couldn't be parsed from Gmail's DOM) that a LATER fetch of the same
            # message sometimes resolves to the real filename instead, making an exact
            # filename match fail even though it's genuinely the same attachment.
            for d in sub_docs:
                d["attachment_index"] = i
            documents.extend(sub_docs)

    email_type = _pick_email_type(documents)
    record = {
        "message_id": message_id,
        "email_type": email_type,
        "doc_types": _distinct_types([d["type"] for d in documents]),
        "type_counts": _type_counts([d["type"] for d in documents if d["source"] == "attachment"]),
        "containers": _merge_containers(*[d.get("containers") for d in documents]),
        "documents": documents,
        "source": "deep",
        "classified_at": datetime.now(timezone.utc).isoformat(),
    }
    with _CACHE_LOCK:
        cache = _load_cache()
        cache[message_id] = record
        _save_cache(cache)

    out = dict(record)
    out["cached"] = False
    return out


def _pick_email_type(documents):
    """Email-level type = the body's type, unless the body is uninformative and a
    confident attachment type exists."""
    if not documents:
        return "Other"
    email_type = documents[0]["type"]
    if email_type in ("Other", "No DOC"):
        best = max(
            (d for d in documents[1:] if d["type"] not in ("Other", "No DOC")),
            key=lambda d: d["confidence"],
            default=None,
        )
        if best:
            email_type = best["type"]
    return email_type


def classify_email(message_id, force=False):
    """DEEP-classify a real Gmail message (body + every attachment) via the read-only
    Gmail API. Raises ``ValueError`` for scraped ('pw_') ids, which the API can't read.
    """
    if not message_id:
        raise ValueError("message_id is required.")
    if message_id.startswith("pw_"):
        raise ValueError(
            "This message lives in the delegated (scraped) mailbox and cannot be read "
            "via the Gmail API. Use the delegated-mailbox deep-scan instead."
        )

    cache = _load_cache()
    if not force and message_id in cache and cache[message_id].get("source") == "deep":
        cached = _backfill_missing_cmr_containers(message_id, dict(cache[message_id]))
        cached["cached"] = True
        return cached

    email_service = EmailService()
    full = email_service.get_email_full(message_id)
    if full is None:
        raise RuntimeError("No Gmail credentials available to retrieve this message.")

    subject = _fetch_subject(email_service, message_id)
    # Only real attachments (inline images referenced by contentId aren't documents).
    raw_attachments = [a for a in full.get("attachments", []) if not a.get("contentId")]

    attachments = []
    for i, att in enumerate(raw_attachments):
        filename = att.get("filename", "")
        mime = _guess_mime(filename)
        data_bytes = None
        if mime.startswith(_LLM_READABLE_PREFIXES) or _is_excel_attachment(filename, mime):
            try:
                data_bytes = email_service.get_attachment(message_id, att["attachmentId"])
            except Exception as e:
                _log_warn(f"Failed to download attachment '{filename}': {e}")
        _cache_attachment_bytes(message_id, i, data_bytes, filename, mime)
        attachments.append({"filename": filename, "mime": mime, "data_bytes": data_bytes,
                            "attachmentId": att.get("attachmentId"), "size": att.get("size", "")})

    result = classify_documents(message_id, subject, full.get("body", ""), attachments, force=force)
    return result


def resolve_deep_classification(message_id, subject="", force=False, label=None):
    """Deep-classify one email regardless of whether it's a real Gmail id or a scraped
    ('pw_') id from a delegated mailbox. ``label`` selects a label override path
    instead of the full classifier - "a-cmr" -> classify_documents_cmr,
    _LCL_ARRIVALS_LABEL ("lcl-arrivals---release") -> classify_documents_lcl."""
    if not message_id.startswith("pw_"):
        # For real Gmail IDs with the a-cmr label, use the CMR override classifier
        # so ALL attachments are extracted (Cmr for the primary, Other for the rest),
        # not just the one the LLM happens to label "Cmr".
        if label == "a-cmr":
            cache = _load_cache()
            if not force and message_id in cache and cache[message_id].get("source") == "deep":
                cached = _backfill_missing_cmr_containers(message_id, dict(cache[message_id]))
                cached["cached"] = True
                return cached
            email_service = EmailService()
            full = email_service.get_email_full(message_id)
            if full is None:
                raise RuntimeError("No Gmail credentials available to retrieve this message.")
            subject_val = _fetch_subject(email_service, message_id)
            raw_attachments = [a for a in full.get("attachments", []) if not a.get("contentId")]
            attachments = []
            for i, att in enumerate(raw_attachments):
                filename = att.get("filename", "")
                mime = _guess_mime(filename)
                data_bytes = None
                if mime.startswith(_LLM_READABLE_PREFIXES) or _is_excel_attachment(filename, mime):
                    try:
                        data_bytes = email_service.get_attachment(message_id, att["attachmentId"])
                    except Exception as e:
                        _log_warn(f"Failed to download attachment '{filename}': {e}")
                attachments.append({"filename": filename, "mime": mime, "data_bytes": data_bytes,
                                    "attachmentId": att.get("attachmentId"), "size": att.get("size", "")})
                _cache_attachment_bytes(message_id, i, data_bytes, filename, mime)
            return classify_documents_cmr(message_id, subject_val, full.get("body", ""), attachments, force=force)
        # Real Gmail IDs with the lcl-arrivals---release label: same shape as the a-cmr
        # branch above, but restricted to the Arrival notice/Delivery order/Other
        # vocabulary via classify_documents_lcl. In practice LCL mail (like a-cmr) is
        # always fetched through the delegated mailbox below ("pw_" ids), but this path
        # is kept for parity/completeness the same way a-cmr's is.
        if label == _LCL_ARRIVALS_LABEL:
            cache = _load_cache()
            if not force and message_id in cache and cache[message_id].get("source") == "deep":
                cached = dict(cache[message_id])
                cached["cached"] = True
                return cached
            email_service = EmailService()
            full = email_service.get_email_full(message_id)
            if full is None:
                raise RuntimeError("No Gmail credentials available to retrieve this message.")
            subject_val = _fetch_subject(email_service, message_id)
            raw_attachments = [a for a in full.get("attachments", []) if not a.get("contentId")]
            attachments = []
            for i, att in enumerate(raw_attachments):
                filename = att.get("filename", "")
                mime = _guess_mime(filename)
                data_bytes = None
                if mime.startswith(_LLM_READABLE_PREFIXES) or _is_excel_attachment(filename, mime):
                    try:
                        data_bytes = email_service.get_attachment(message_id, att["attachmentId"])
                    except Exception as e:
                        _log_warn(f"Failed to download attachment '{filename}': {e}")
                attachments.append({"filename": filename, "mime": mime, "data_bytes": data_bytes,
                                    "attachmentId": att.get("attachmentId"), "size": att.get("size", "")})
                _cache_attachment_bytes(message_id, i, data_bytes, filename, mime)
            return classify_documents_lcl(message_id, subject_val, full.get("body", ""), attachments, force=force)
        return classify_email(message_id, force=force)

    # classify_documents() below already has this exact cache check built in - but
    # only AFTER paying for a full browser round-trip (open the email, wait for the
    # body to render, download every attachment) to build the arguments it needs.
    # Checking here first means an already-classified message returns instantly
    # instead of re-opening it in the automation browser on every call - previously,
    # repeated classification requests for the same message (e.g. the dashboard
    # re-checking a job, or a user re-selecting the same email) each cost 45-90+
    # seconds and repeatedly clicking the same Gmail row back-to-back like that was
    # itself a source of flakiness (the tab's "return to list" navigation racing with
    # the next request's row click).
    if not force:
        cache = _load_cache()
        if message_id in cache and cache[message_id].get("source") == "deep":
            cached = _backfill_missing_cmr_containers(message_id, dict(cache[message_id]))
            cached["cached"] = True
            return cached

    raw_id = message_id[len("pw_"):]
    params = urllib.parse.urlencode({"id": raw_id, "subject": subject})
    url = f"http://127.0.0.1:40005/get_documents?{params}"
    # Matches open_gmail.py's /get_documents internal wait (widened to 150s - a
    # 2-attachment classification measured at ~90s end-to-end, right at the old 90s
    # ceiling here, causing an otherwise-successful multi-attachment read to be thrown
    # away and silently replaced with the single-attachment metadata guess).
    with urllib.request.urlopen(url, timeout=160) as response:
        payload = json.loads(response.read().decode("utf-8"))

    attachments = []
    for i, att in enumerate(payload.get("attachments", [])):
        b64 = att.get("data_b64")
        data_bytes = None
        if b64:
            try:
                data_bytes = base64.b64decode(b64)
            except Exception:
                data_bytes = None
        filename = att.get("filename", "")
        mime = att.get("mime", "")
        attachments.append({"filename": filename, "mime": mime, "data_bytes": data_bytes})
        # Persist the bytes we just paid a full browser round-trip for, so the later
        # Shypple upload step (fetch_document_bytes) can read them straight off disk
        # instead of re-opening this same mail a second time just to get bytes it
        # already downloaded once, right here, minutes earlier.
        _cache_attachment_bytes(message_id, i, data_bytes, filename, mime)

    if label == "a-cmr":
        return classify_documents_cmr(message_id, subject, payload.get("body", ""), attachments, force=force)
    if label == _LCL_ARRIVALS_LABEL:
        return classify_documents_lcl(message_id, subject, payload.get("body", ""), attachments, force=force)
    return classify_documents(message_id, subject, payload.get("body", ""), attachments, force=force)


def classify_documents_cmr(message_id, subject, body_html, attachments, force=False):
    """CMR-label override of classify_documents. Every mail in label:a-cmr carries a
    CMR document as its primary attachment - but we should classify ALL attachments properly.

      - First, identify which attachment is the CMR document using filename hints and content analysis.
      - For the CMR attachment, classify as "Cmr".
      - For all other attachments, use the regular classifier to determine their actual type
        (not just mark them as "Other").
      - Excel/body-only entries fall through to the base classifier unchanged.
      - All confirmation gates (awaiting_upload_confirmation, awaiting_submit_
        confirmation, etc.) are unaffected - this only decides the TYPE label, nothing
        else in the pipeline changes.

    Implemented as a thin wrapper: builds the attachments list with pre-assigned types
    and calls classify_documents with those overrides baked in, so caching, container
    extraction, and every downstream consumer work exactly as before."""
    # 1. Check filename hints first to identify CMR document
    cmr_index = None
    for i, att in enumerate(attachments or []):
        if _is_excel_attachment(att.get("filename", ""), att.get("mime", "")):
            continue
        if filename_hint(att.get("filename", "")) == "Cmr":
            cmr_index = i
            break

    # 2. Check PDF text content to accurately identify CMR vs Status Update / Other
    if cmr_index is None:
        other_indices = set()
        for i, att in enumerate(attachments or []):
            if _is_excel_attachment(att.get("filename", ""), att.get("mime", "")):
                continue
            data_bytes = att.get("data_bytes")
            if data_bytes and (att.get("mime") == "application/pdf" or att.get("filename", "").lower().endswith(".pdf")):
                page_texts = _extract_pdf_page_texts(data_bytes)
                text = " ".join(page_texts).lower()
                is_status_update = bool(re.search(r"status\s*update|mededeling|vooraanmelding|pakket|packing\s*list|invoice|factura", text))
                is_cmr_text = bool(re.search(r"\bcmr\b|vrachtbrief|consignment\s*note|lettre\s*de\s*voiture|internationaler\s*frachtbrief", text))
                if is_status_update:
                    other_indices.add(i)
                if is_cmr_text and not is_status_update and cmr_index is None:
                    cmr_index = i

        # 3. Default to the first non-Excel attachment that is NOT a Status Update / Other
        if cmr_index is None:
            for i, att in enumerate(attachments or []):
                if not _is_excel_attachment(att.get("filename", ""), att.get("mime", "")) and i not in other_indices:
                    cmr_index = i
                    break

    # Fallback to index 0 if all attachments were non-Excel
    if cmr_index is None:
        for i, att in enumerate(attachments or []):
            if not _is_excel_attachment(att.get("filename", ""), att.get("mime", "")):
                cmr_index = i
                break

    # For CMR attachments, we override to "Cmr", for others we let the regular classifier determine the type
    overridden = []
    for i, att in enumerate(attachments or []):
        if _is_excel_attachment(att.get("filename", ""), att.get("mime", "")):
            overridden.append(att)  # unchanged - base classifier handles Excel
            continue
        patched = dict(att)
        patched["_cmr_type_override"] = "Cmr" if i == cmr_index else "Other"
        overridden.append(patched)

    return _classify_documents_with_overrides(message_id, subject, body_html, overridden, force=force)


def classify_documents_lcl(message_id, subject, body_html, attachments, force=False):
    """lcl-arrivals---release label override of classify_documents - the deep-read
    counterpart of classify_email_meta's _LCL_ARRIVALS_LABEL branch above, and the
    same parallel classify_documents_cmr already draws for a-cmr's Cmr/Other split.
    This label's document-type universe is restricted to "Arrival notice"/
    "Delivery order"/"Other" only - the generic classifier's full CMR taxonomy
    (Packing List, Final master bill of lading, etc.) doesn't apply here and was
    exactly what caused unrelated types to show up for LCL mail.

    Reuses classify_lcl_arrival_email's already-tuned regex rules (the same ones that
    route these mails to their processing handler in process_lcl_arrival_job) rather
    than a second, separate guesser - body_html is reduced to plain text to stand in
    for that function's usual "snippet" input. Per the operator's explicit confirmation,
    the actual Arrival notice/Delivery order PDFs print that exact wording in their own
    top section, so EACH real attachment's own extracted text is folded into its OWN
    classification check below (not just a single "primary" attachment's) - this is
    what lets a document whose own printed heading says "ARRIVAL NOTICE"/"DELIVERY
    ORDER" be trusted even when the subject line alone is ambiguous (e.g. a generic
    forward), and, per the operator's explicit example, what lets a SECOND real
    document in the same mail (e.g. a carrier's own "Release" PDF alongside a
    "Delivery order" PDF) also correctly resolve to "Delivery order" on its own merit
    instead of being forced to "Other" just for not being attachment index 0.
    classify_lcl_arrival_email's existing hard overrides (status update / transfer
    notification) still apply first for every attachment, since they also search this
    same per-attachment combined text."""
    body_text = _html_to_text(body_html or "")[:2000]

    per_att_types = []
    overridden = []
    for i, att in enumerate(attachments or []):
        if _is_excel_attachment(att.get("filename", ""), att.get("mime", "")):
            overridden.append(att)
            per_att_types.append(None)
            continue

        doc_text = ""
        data_bytes = att.get("data_bytes")
        mime = att.get("mime") or _guess_mime(att.get("filename", ""))
        if data_bytes and mime == "application/pdf":
            doc_text = " ".join(_extract_pdf_page_texts(data_bytes))[:4000]

        att_mail_type = classify_lcl_arrival_email({
            "subject": subject,
            "snippet": (body_text + " " + doc_text).strip(),
            "attachmentNames": [att.get("filename", "")],
        })
        att_type = _LCL_MAIL_TYPE_TO_DOC_TYPE.get(att_mail_type, "Other")
        per_att_types.append(att_type)

        patched = dict(att)
        patched["_cmr_type_override"] = att_type
        overridden.append(patched)

    # This mail's own overall type (used e.g. by lcl_arrivals_process's mail_type
    # "unknown" fallback) is whichever real (non-"Other") type the FIRST attachment
    # resolved to - matches the pre-existing "primary attachment drives the mail type"
    # behavior; only the per-OTHER-attachment forcing to "Other" regardless of its own
    # content has been removed above.
    resolved_type = next((t for t in per_att_types if t and t != "Other"), "Other")
    mail_type = next((k for k, v in _LCL_MAIL_TYPE_TO_DOC_TYPE.items() if v == resolved_type), "unknown")

    return _classify_documents_with_overrides(
        message_id, subject, body_html, overridden, force=force,
        override_method="lcl-label-override",
        override_reason=f"Classified by lcl-arrivals label rule (mail_type={mail_type}).",
        override_confident_types=("Arrival notice", "Delivery order"),
    )


def _extract_containers_via_llm(gemini, data_bytes, mime, filename):
    """Read ISO-6346 container number(s) directly off a document's actual bytes via
    Gemini multimodal. Fallback for exactly the case _extract_pdf_page_texts's own
    docstring warns about but the CMR-override path (below) never actually used: a
    scanned or handwritten PDF/image with no extractable text layer - very common for
    a CMR consignment note, which is normally a hand-filled form - where the
    deterministic regex pass over the (empty) text layer finds nothing even though a
    human reading the page can see the container number right there. Best-effort only;
    returns [] on any failure or on a non-multimodal-readable file, never raises."""
    if not data_bytes or not (mime or "").startswith(_LLM_READABLE_PREFIXES) or len(data_bytes) > 18 * 1024 * 1024:
        return []
    prompt = (
        "This is a scanned freight/logistics document (e.g. a CMR consignment note). "
        "Read EVERY ISO-6346 container number printed OR handwritten on it - 4 letters "
        "(a 3-letter owner code + U/J/Z category code) followed by 7 digits, e.g. "
        "TEMU9681744 or CAIU7160568. Handwritten numbers can be messy - read carefully. "
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"containers": ["..."]}'
    )
    try:
        parsed = _parse_llm_json(gemini.generate_multimodal(prompt, [(mime, data_bytes)]))
        if parsed:
            return _merge_containers(parsed.get("containers"))
    except Exception as e:
        _log_warn(f"Container-only LLM read failed for '{filename}': {e}")
    return []


def _backfill_missing_cmr_containers(message_id, record):
    """Self-heal a cached "deep" a-cmr record that has zero containers anywhere -
    e.g. one classified before _extract_containers_via_llm existed, so its Cmr
    document's containers were permanently stuck at [] the moment it got cached
    (a "deep" cache hit is normally trusted as-is and never re-processed). Re-reads
    the Cmr-typed attachment's bytes from the on-disk attachment cache (already saved
    by every deep classification via _cache_attachment_bytes - no Gmail re-open
    needed) and runs the same LLM container fallback a fresh classification would now
    use, patching the cache in place if it finds something.

    Runs at most once per record: marks "containers_backfill_attempted" regardless of
    outcome, so a mail that genuinely has no container number printed anywhere isn't
    re-sent to Gemini on every single future lookup. Best-effort - any failure here
    just leaves the record as it was, never raises."""
    if record.get("containers") or record.get("containers_backfill_attempted"):
        return record

    cmr_doc = next(
        (d for d in record.get("documents", [])
         if d.get("source") == "attachment" and d.get("type") == "Cmr"),
        None,
    )
    if cmr_doc is None or cmr_doc.get("attachment_index") is None:
        return record

    try:
        data_bytes, cached_mime, cached_filename = _load_cached_attachment(message_id, cmr_doc["attachment_index"])
        if not data_bytes:
            return record
        gemini = GeminiClient()
        containers = _extract_containers_via_llm(
            gemini, data_bytes,
            cached_mime or "", cached_filename or cmr_doc.get("filename", ""),
        )
    except Exception as e:
        _log_warn(f"Container backfill failed for {message_id}: {e}")
        containers = []

    with _CACHE_LOCK:
        cache = _load_cache()
        cached_rec = cache.get(message_id)
        if cached_rec is None:
            return record
        cached_rec["containers_backfill_attempted"] = True
        if containers:
            cached_rec["containers"] = _merge_containers(cached_rec.get("containers"), containers)
            for d in cached_rec.get("documents", []):
                if d.get("source") == "attachment" and d.get("type") == "Cmr":
                    d["containers"] = _merge_containers(d.get("containers"), containers)
        cache[message_id] = cached_rec
        _save_cache(cache)
        return dict(cached_rec)


def _classify_documents_with_overrides(message_id, subject, body_html, attachments, force=False,
                                        override_method="cmr-label-override",
                                        override_reason="Classified by a-cmr label rule.",
                                        override_confident_types=("Cmr",)):
    """classify_documents variant that respects a ``_cmr_type_override`` key on each
    attachment dict. When present, the attachment's type is set directly (skipping the
    LLM guesser for that slot) while containers are still extracted from the bytes.
    All other behaviour - caching, body classification, container merging - is
    identical to classify_documents.

    ``override_method``/``override_reason``/``override_confident_types`` let a
    different label override (classify_documents_lcl) reuse this exact machinery with
    its own type vocabulary and audit text instead of the CMR-flavoured defaults -
    ``override_confident_types`` is which override values count as a confident,
    deterministic match (confidence 1.0) rather than the "Other" fallback (0.0)."""
    cache = _load_cache()
    if not force and message_id in cache and cache[message_id].get("source") == "deep":
        cached = _backfill_missing_cmr_containers(message_id, dict(cache[message_id]))
        cached["cached"] = True
        return cached

    body_text = _html_to_text(body_html or "")
    gemini = GeminiClient()

    documents = [_classify_body(gemini, subject, body_text)]
    for i, att in enumerate(attachments or []):
        filename = att.get("filename", "") or ""
        mime = att.get("mime") or _guess_mime(filename)
        override_type = att.get("_cmr_type_override")
        if override_type:
            # Extract containers from bytes deterministically; skip the LLM type guess.
            data_bytes = att.get("data_bytes")
            page_texts = _extract_pdf_page_texts(data_bytes) if (data_bytes and mime == "application/pdf") else []
            containers = _merge_containers(
                find_container_numbers(" ".join(page_texts)),
                find_container_numbers(filename),
            )
            if not containers:
                # No text layer (scanned/handwritten CMR) and nothing in the filename -
                # the deterministic pass above has nothing left to work with even though
                # the container number is plainly visible on the page. Ask Gemini to
                # read it directly off the image/PDF bytes instead of leaving this
                # document's containers empty.
                containers = _extract_containers_via_llm(gemini, data_bytes, mime, filename)
            sub_docs = [{
                "source": "attachment",
                "filename": filename,
                "type": override_type,
                "pages": "all",
                "confidence": 1.0 if override_type in override_confident_types else 0.0,
                "method": override_method,
                "reason": override_reason,
                "containers": containers,
                "attachment_index": i,
            }]
        else:
            sub_docs = _classify_attachment(gemini, subject, body_text, filename, att.get("data_bytes"), mime)
            for d in sub_docs:
                d["attachment_index"] = i
        documents.extend(sub_docs)

    email_type = _pick_email_type(documents)
    record = {
        "message_id": message_id,
        "email_type": email_type,
        "doc_types": _distinct_types([d["type"] for d in documents]),
        "type_counts": _type_counts([d["type"] for d in documents if d["source"] == "attachment"]),
        "containers": _merge_containers(*[d.get("containers") for d in documents]),
        "documents": documents,
        "source": "deep",
        "classified_at": datetime.now(timezone.utc).isoformat(),
    }
    with _CACHE_LOCK:
        cache = _load_cache()
        cache[message_id] = record
        _save_cache(cache)

    out = dict(record)
    out["cached"] = False
    return out


def find_document_source(message_id, doc_type, attachment_index=None):
    """Which attachment (if any) in this message's cached classification was
    classified as ``doc_type``. Only matches attachment-sourced entries."""
    cache = _load_cache()
    record = cache.get(message_id)
    if not record:
        return None

    # First pass: try exact attachment_index match if provided
    if attachment_index is not None:
        for doc in record.get("documents", []):
            if doc.get("source") == "attachment" and doc.get("attachment_index") == attachment_index:
                return doc

    # Second pass: match by doc_type
    for doc in record.get("documents", []):
        if doc.get("source") == "attachment" and doc.get("type") == doc_type:
            return doc

    # Third pass: check raw attachments list if present
    for att in record.get("attachments", []):
        if att.get("type") == doc_type:
            return att

    return None


def find_all_document_sources(message_id, doc_type):
    """Every attachment-sourced document in this message's cached classification
    matching ``doc_type`` - not just the first (see find_document_source above), for
    the case where a single mail genuinely carries more than one document of the same
    type (e.g. a Delivery Order mail with both a carrier's own "Delivery order" PDF and
    a separate "Release" document that also resolves to "Delivery order" on its own
    merit - see classify_documents_lcl's per-attachment classification). Ordered by
    attachment_index. Returns [] if nothing is cached yet or nothing matches."""
    cache = _load_cache()
    record = cache.get(message_id)
    if not record:
        return []
    docs = [
        d for d in record.get("documents", [])
        if d.get("source") == "attachment" and d.get("type") == doc_type
    ]
    docs.sort(key=lambda d: d.get("attachment_index") if d.get("attachment_index") is not None else 0)
    return docs


def fetch_document_bytes(message_id, doc_type, subject="", attachment_index=None, label="a-cmr"):
    """Re-download the actual file bytes for the attachment classified as ``doc_type``
    in this message - used by the Operations Process automation both to compare an
    already-uploaded Shypple document against the email's version, and to supply the
    bytes for an actual upload. ``subject`` (if known) lets the delegated-mailbox path
    below find the message by searching Gmail when its row has scrolled out of the
    currently-loaded list view - without it, a message that's still real but no longer
    on screen fails outright. ``label`` is passed through to the force re-classify retry
    below - defaults to "a-cmr" since most callers are the CMR flow. Callers for
    lcl-arrivals mail (e.g. lcl_arrivals_process in operations_api.py) MUST pass
    label=_LCL_ARRIVALS_LABEL ("lcl-arrivals---release") instead, or that retry would
    fall through to the generic full-CMR-taxonomy classifier - which doesn't just
    mislabel the type, it can make THIS function find no document at all for
    doc_type="Delivery order"/"Arrival notice" if the generic guess picked a different
    type, permanently corrupting the cached classification either way. Returns
    ``(data_bytes, mime, filename)``, or ``(None, None, None)`` if no matching
    attachment is on record or it can't be re-fetched."""
    doc = find_document_source(message_id, doc_type, attachment_index=attachment_index)
    # A cached "deep" record with no attachment-sourced entry for this type usually
    # means the attachment-finding DOM scan missed on an EARLIER run (e.g. a timing
    # race - see open_gmail.py's fetch_email_body retry). A record classified before
    # "attachment_index" existed is the same kind of staleness - it can only be
    # matched by filename, which is unreliable (see below). Either way, since "deep" is
    # normally treated as already-complete and never retried, that staleness would
    # otherwise be permanent - force exactly one fresh re-classification before giving
    # up, in case this was a transient scan failure rather than the attachment
    # genuinely not existing.
    needs_retry = (not doc) or (message_id.startswith("pw_") and doc.get("attachment_index") is None)
    if needs_retry and message_id.startswith("pw_"):
        try:
            resolve_deep_classification(message_id, subject=subject, force=True, label=label)
        except Exception as e:
            _log_warn(f"fetch_document_bytes: force re-classify retry failed for {message_id}: {e}")
        doc = find_document_source(message_id, doc_type, attachment_index=attachment_index)
    if not doc:
        return None, None, None
    filename = doc.get("filename", "")
    attachment_index = doc.get("attachment_index")

    # Serve straight from disk if classification already downloaded and cached these
    # exact bytes (see _cache_attachment_bytes) - this is the common case once a mail
    # has been through "Find Document Types" or the Operations Process review, and
    # skips re-opening the mail in the automation browser a second time just to fetch
    # bytes it already fetched once, minutes earlier.
    cached_bytes, cached_mime, cached_filename = _load_cached_attachment(message_id, attachment_index)
    if cached_bytes is not None:
        return cached_bytes, cached_mime or _guess_mime(cached_filename or filename), cached_filename or filename

    if not message_id.startswith("pw_"):
        email_service = EmailService()
        full = email_service.get_email_full(message_id)
        if not full:
            return None, None, None
        for att in full.get("attachments", []):
            if att.get("filename") == filename:
                try:
                    data_bytes = email_service.get_attachment(message_id, att["attachmentId"])
                except Exception as e:
                    _log_warn(f"fetch_document_bytes: could not download '{filename}': {e}")
                    return None, None, None
                _cache_attachment_bytes(message_id, attachment_index, data_bytes, filename, _guess_mime(filename))
                return data_bytes, _guess_mime(filename), filename
        return None, None, None

    # Scraped mailbox - proxy to the Gmail automation's /get_documents, which already
    # fetches every attachment's bytes for that message in one call.
    raw_id = message_id[len("pw_"):]
    try:
        params = urllib.parse.urlencode({"id": raw_id, "subject": subject})
        url = f"http://127.0.0.1:40005/get_documents?{params}"
        # Matches open_gmail.py's /get_documents internal wait (widened to 150s - a
        # 2-attachment classification measured at ~90s end-to-end, right at the old 90s
        # ceiling here, causing an otherwise-successful multi-attachment read to be thrown
        # away and silently replaced with the single-attachment metadata guess).
        with urllib.request.urlopen(url, timeout=160) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as e:
        _log_warn(f"fetch_document_bytes: could not reach the delegated-mailbox browser: {e}")
        return None, None, None

    fresh_attachments = payload.get("attachments", [])

    # Match by POSITION first - the filename recorded at classify time can be a
    # generic fallback ("Attachment 1") that a later fetch resolves to the real
    # filename instead, so an exact filename match can miss the very same attachment
    # (see classify_documents' attachment_index comment). Position is stable because
    # both fetches scan the same message's attachment anchors in the same DOM order.
    match = None
    if attachment_index is not None and 0 <= attachment_index < len(fresh_attachments):
        match = fresh_attachments[attachment_index]

    # Fall back to a filename match for cache entries with no attachment_index at all
    # (shouldn't normally happen after the retry above, but cheap insurance) or if the
    # attachment count genuinely changed between fetches.
    if match is None:
        for att in fresh_attachments:
            if att.get("filename") == filename:
                match = att
                break

    if match is None:
        return None, None, None
    b64 = match.get("data_b64")
    if not b64:
        return None, None, None
    real_filename = match.get("filename") or filename
    try:
        data_bytes = base64.b64decode(b64)
    except Exception:
        return None, None, None
    real_mime = match.get("mime") or _guess_mime(real_filename)
    _cache_attachment_bytes(message_id, attachment_index, data_bytes, real_filename, real_mime)
    return data_bytes, real_mime, real_filename


def compare_document_versions(email_bytes, email_mime, other_bytes, other_mime, doc_type):
    """Ask Gemini whether two versions of a document classified as the SAME type are
    actually the same document (same shipment/dates/parties/amounts) or a materially
    different one - used before trusting an already-uploaded Shypple document just
    because its type label matches, per the operator's explicit request to verify
    deeply rather than assume, and to require a human to actually see and approve any
    real difference rather than a one-line summary. Returns {"same": bool|None,
    "reason": str, "differences": [...]}` - "same" of None means the comparison itself
    failed (couldn't read one or both files), not a verdict. "differences" is a
    field-by-field list (only populated when same is False) of
    {"field", "email_value", "shypple_value"}, so the dashboard can show the operator
    exactly what differs instead of just Gemini's prose summary."""
    if not email_bytes or not other_bytes:
        return {"same": None, "reason": "One or both documents were unavailable to compare.", "differences": []}
    if not (email_mime or "").startswith(_LLM_READABLE_PREFIXES) or not (other_mime or "").startswith(_LLM_READABLE_PREFIXES):
        return {"same": None, "reason": "One or both documents aren't in a format Gemini can read inline.", "differences": []}

    gemini = GeminiClient()
    prompt = (
        "You are comparing two versions of a logistics/trade document that were both "
        f'classified as "{doc_type}". The FIRST file is from the tracked email; the '
        "SECOND is what's already uploaded on the shipment record. Read BOTH carefully "
        "and compare the key fields: shipment/BL/booking reference, dates (ETA/ETD/issue "
        "date), parties/addresses, container number(s), quantities, and amounts - "
        "formatting/scan-quality differences (date format, whitespace, line breaks) are "
        "NOT real differences and must not be reported as one.\n\n"
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"same": true/false, "reason": "<one sentence summary>", '
        '"differences": [{"field": "<field name>", "email_value": "<value in the email\'s copy>", '
        '"shypple_value": "<value in Shypple\'s copy>"}]}\n\n'
        'If same is true, "differences" must be an empty array. If same is false, '
        '"differences" must list every field whose value genuinely differs (not just '
        "reformatted) between the two."
    )
    # Two attempts: json_mode constrains Gemini to valid JSON directly (rather than
    # relying on prompt wording + a regex extractor), but a single retry still catches
    # the rare truncated/empty response without doubling normal-case latency by default.
    last_reason = "Comparison failed - Gemini did not return a usable verdict."
    for attempt in range(2):
        try:
            raw = gemini.generate_multimodal(
                prompt, [(email_mime, email_bytes), (other_mime, other_bytes)], json_mode=True,
            )
        except Exception as e:
            last_reason = f"Gemini call failed: {e}"
            _log_warn(f"compare_document_versions LLM call failed (attempt {attempt + 1}): {e}")
            continue

        parsed = _parse_llm_json(raw)
        if not parsed:
            last_reason = "Gemini's response wasn't valid JSON."
            _log_warn(f"compare_document_versions got non-JSON response (attempt {attempt + 1}): {raw[:300]!r}")
            continue
        if "same" not in parsed:
            last_reason = "Gemini's response was missing a same/different verdict."
            _log_warn(f"compare_document_versions response missing 'same' (attempt {attempt + 1}): {raw[:300]!r}")
            continue

        raw_diffs = parsed.get("differences")
        differences = [
            {
                "field": str(d.get("field", ""))[:80],
                "email_value": str(d.get("email_value", ""))[:200],
                "shypple_value": str(d.get("shypple_value", ""))[:200],
            }
            for d in (raw_diffs if isinstance(raw_diffs, list) else [])
            if isinstance(d, dict)
        ]
        # Gemini's own "same" verdict and its itemized differences can disagree - it
        # said "different" but then listed a field with the IDENTICAL value on both
        # sides (observed live: "Container number(s): HLBU6079748" vs "HLBU6079748").
        # Drop any entry that isn't an actual difference, and if nothing genuine
        # survives, trust that evidence over the verdict rather than showing the
        # operator a misleading "differs" table with matching values in both columns.
        real_differences = [
            d for d in differences
            if d["email_value"].strip().casefold() != d["shypple_value"].strip().casefold()
        ]
        same = bool(parsed["same"])
        reason = str(parsed.get("reason", ""))[:400]
        if not same and not real_differences:
            same = True
            reason = (reason + " (no field-level difference could be confirmed on review - treating as a match.)").strip()
        return {
            "same": same,
            "reason": reason,
            "differences": real_differences,
        }
    return {"same": None, "reason": last_reason, "differences": []}


# --------------------------------------------------------------------------- #
# META path - READ-SAFE, list-view metadata only (backs the bulk button).
# --------------------------------------------------------------------------- #

def _build_meta_prompt(subject, snippet, sender, attachment_names, has_attachment=False):
    lines = [
        f"Sender: {sender or '(unknown)'}",
        f"Subject: {subject or '(none)'}",
        f"Preview snippet: {snippet or '(none)'}",
    ]
    if attachment_names:
        lines.append("Attachment file names: " + "; ".join(attachment_names))
    elif has_attachment:
        # The list-row scrape confirmed a real attachment chip exists, but couldn't read
        # its filename (title/inner text extraction miss) - tell the model explicitly so
        # it doesn't fall back to "No DOC" just because the name is unknown.
        lines.append(
            "Attachment file names: (unknown - this mail DOES have at least one real "
            "attachment, its filename just couldn't be read from the list view)"
        )
    context = "\n".join(lines)

    attachment_note = (
        "\nNote: this mail is CONFIRMED to carry a real attachment even though its "
        "filename is unknown - do not answer \"No DOC\" for that reason alone. Infer the "
        "type from the sender/subject/preview if possible, otherwise answer \"Other\".\n"
        if has_attachment and not attachment_names else ""
    )

    return (
        "You classify logistics / import trade documents for a freight-forwarding workflow.\n"
        "You are given only an email's list-view metadata (sender, subject, a short preview,\n"
        "and attachment file names) - you do NOT have the full body or the attachment bytes.\n\n"
        f"Allowed types (copy verbatim, exact spelling):\n{_numbered_types()}\n\n"
        f"{_reference_block()}\n\n"
        f"{_few_shot_block()}\n\n"
        "HOW TO DECIDE (in this order):\n"
        "1. KEYWORD: look for an explicit document-type keyword in the attachment file\n"
        "   names or the subject, and map it through the reference table.\n"
        "2. If there is no such keyword, infer the type from the sender + subject + preview.\n\n"
        f"{_GUARDRAIL}"
        f"{attachment_note}\n"
        "Also extract EVERY ISO-6346 container number in the subject/preview/file names\n"
        "(4 letters + 7 digits, e.g. TEMU9681744) into \"containers\" - there may be more\n"
        "than one, list all of them.\n\n"
        f"{context}\n\n"
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"type": "<one exact value>", "confidence": <0.0-1.0>, "reason": "<short>", "containers": ["..."]}'
    )


def extract_sf_number(text):
    """Extract SF number like SF169508 from email subject, body or text."""
    if not text:
        return None
    m = re.search(r"\bSF\d+\b", text, re.IGNORECASE)
    return m.group(0).upper() if m else None


def _normalize_date_str(raw):
    """Turn a raw date substring (YYYY-MM-DD, DD-MM-YYYY / DD/MM/YYYY / DD.MM.YYYY, DD Mon YYYY, or DD/MM) into
    YYYY-MM-DD. Returns None if raw doesn't look like these shapes."""
    if not raw:
        return None
    raw = raw.strip()
    
    # Check YYYY-MM-DD or YYYY/MM/DD or YYYY.MM.DD
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        
    # Check DD-MM-YYYY or DD/MM/YYYY or DD.MM.YYYY
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", raw)
    if m:
        day, month, year = m.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    # Check DD-MM-YY or DD/MM/YY or DD.MM.YY
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2})$", raw)
    if m:
        day, month, yy = m.groups()
        year = 2000 + int(yy)
        return f"{year}-{int(month):02d}-{int(day):02d}"

    # Check DD Mon YYYY (e.g. 15 Aug 2026 or 15-Aug-2026)
    m = re.match(r"^(\d{1,2})[\s\-/.]?([A-Za-z]{3,9})[\s\-/.]?(\d{4})$", raw)
    if m:
        day, mon_str, year = m.groups()
        try:
            from datetime import datetime
            dt = datetime.strptime(f"{day} {mon_str[:3]} {year}", "%d %b %Y")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
        
    # Check DD-MM or D-MM (partial date, e.g. 4/08)
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})$", raw)
    if m:
        day, month = m.groups()
        from datetime import datetime
        year = datetime.now().year
        return f"{year}-{int(month):02d}-{int(day):02d}"
        
    return None


# Labels under which expected devanning date, available date at CFS, delay date, or
# unpacked/stripping date can be printed in email text or attached documents.
# Strictly requires an explicit label so unrelated dates (such as vessel ETA, ATA,
# or document issue date) are NOT mistakenly grabbed when no devanning date exists.
_DEVANNING_DATE_LABEL_RE = re.compile(
    r"(?:"
    r"expected\s+devanning\s+date|devan(?:ning|ing|aing|ing)?\s+date|devan(?:ning|ing|aing|ing)?"
    r"|available\s+date\s+(?:at\s+cfs)?|cfs\s+available\s+date|cfs\s+avail(?:able|\.)?\s+date|availability\s+date|available\s+at\s+cfs|available\s+from|available\s+after|avail\.?\s+date|date\s+of\s+availability"
    r"|delaye?d?\s+(?:date|to|until)|new\s+devan(?:ning|ing|aing|ing)?\s+date|revised\s+(?:devan(?:ning|ing|aing|ing)?\s+)?date|new\s+available\s+date|revised\s+availab(?:ility|le)\s+date|postponed\s+(?:to|until)"
    r"|unpacked?\s+date|stripping\s+date|storage\s+start(?:ing)?|free\s+time\s+until|pickup\s+date|cfs\s+date"
    r")"
    r"[\s\:\=]*[^\r\n]{0,50}?\b([0-9]{1,4}[-/.\s][0-9A-Za-z]{1,9}(?:[-/.\s][0-9]{2,4})?)\b",
    re.IGNORECASE,
)


_INVALID_CUSTOMS_WORDS_RE = re.compile(
    r"\b(?:"
    r"vessel|voyage|vsl|vyg|voy|ship|feeder|flag|imo|mmsi"
    r"|n/?a|tbd|none|nil|null|unknown|pending|unavailable|not\s+available|available\s+after|after\s+devanning|available"
    r"|eta|etd|ata|atd|pod|pol|port|discharge|cfs|warehouse|address"
    r"|bill\s+of\s+lading|b/?l|container|seal|weight|volume|packages|pcs|gross|net"
    r"|maersk|msc|cma|cgm|cosco|evergreen|hapag|lloyd|oocl|yang\s*ming|wan\s*hai|hyundai|hmm|zim|pil|kmc|sitc|whl|dfds|unifeeder"
    r"|isabella|marco|polo|given|ever|apus|express|spirit|bridge|haven|harbor|bay|star|ocean|sea|pacific|atlantic"
    r")\b",
    re.IGNORECASE,
)


def _is_valid_customs_number(candidate):
    """Sanity-check a preceding customs number candidate. Returns False if candidate
    is blank, contains no digits, is a date, or contains vessel/voyage/status words
    that flattened PDF table extraction frequently grabs from adjacent cells when
    the customs number field is genuinely blank."""
    if not candidate:
        return False
    candidate = candidate.strip()
    if not re.search(r"\d", candidate):
        return False
    if _INVALID_CUSTOMS_WORDS_RE.search(candidate):
        return False
    if re.match(r"^\d{4}[-/.]\d{2}[-/.]\d{2}$", candidate) or re.match(r"^\d{2}[-/.]\d{2}[-/.]\d{4}$", candidate):
        return False
    if len(candidate) < 4 or len(candidate) > 35:
        return False
    return True


def extract_lcl_arrival_data(text):
    """Extract data fields required for LCL Arrivals:
    - container_number (ISO 6346)
    - devanning_date / available date at cfs (YYYY-MM-DD)
    - customs_number (preceding customs number, e.g., 641761FPS-01)
    - cfs_address / warehouse (e.g. CTG Logistics)
    """
    if not text:
        return {}

    containers = find_container_numbers(text)
    container_number = containers[0] if containers else None

    # Date extraction: ONLY extract a date if it is explicitly labeled as a
    # devanning, available, or delay date via _DEVANNING_DATE_LABEL_RE.
    # We deliberately do NOT fall back to bare/unlabeled date scans in the document
    # because that would mistakenly grab unrelated dates like vessel ETA or issue date.
    label_match = _DEVANNING_DATE_LABEL_RE.search(text)
    devanning_date = _normalize_date_str(label_match.group(1)) if label_match else None

    # Customs number match (e.g. 641761FPS-01 or Customs Number label).
    customs_match = re.search(
        r"(?:customs\s*number|previous\s*customs?\s*number|preceding\s*customs?\s*number|customs\s*no\.?|customs\s*ref(?:erence)?)[:\s]*([A-Z0-9\-_/]{3,35})",
        text, re.IGNORECASE
    )
    customs_number = None
    if customs_match:
        candidate = customs_match.group(1).strip()
        if _is_valid_customs_number(candidate):
            customs_number = candidate

    if not customs_number:
        # Check standard preceding customs number formats (e.g. 641761FPS-01)
        for match in re.finditer(r"\b(\d{5,}[A-Z0-9\-_]{2,})\b", text):
            cand = match.group(1).strip()
            if _is_valid_customs_number(cand):
                customs_number = cand
                break

    # CFS Address / Warehouse - only the FIRST LINE of the address (e.g. "VLS
    # BELGIUM"), per the operator's explicit request: the Shypple side only ever needs
    # the first 3 letters of this (see edit_preceding_customs_and_cfs's prefix match
    # just below), and the warehouse block on a real Arrival Notice runs on for
    # several more lines of operational detail (opening hours, reception email,
    # entrepot number, etc.) that must NOT end up in this field. The old unbounded
    # `[^\r\n]+` swallowed that whole block on a real document whose PDF text has no
    # real newlines between address lines - some of these are HTML-to-PDF exports that
    # bake "<br>" in as literal visible text instead of an actual line break (confirmed
    # live: "Warehouse:VLS BELGIUM<br>ROMEYNSWEEL 8<br>HAVENNUMMER: ..."). Excluding
    # "<" from the captured class stops at that literal tag; the 80-char cap is a
    # last-resort safety net for a genuinely run-on block with neither a real newline
    # nor a "<br>" anywhere.
    cfs_match = re.search(r"(?:cfs\s*address|warehouse|discharge\s*cfs)[:\s]*([^\r\n<]{1,80})", text, re.IGNORECASE)
    cfs_address = cfs_match.group(1).strip() if cfs_match else None

    return {
        "container_number": container_number,
        "devanning_date": devanning_date,
        "customs_number": customs_number,
        "cfs_address": cfs_address,
    }


def extract_lcl_fields_via_llm(gemini, data_bytes, mime, filename):
    """Read the LCL Arrivals fields (plus the SF number) directly off a document's
    actual bytes via Gemini multimodal - fallback for when extract_lcl_arrival_data's
    deterministic regex pass over the PDF's text layer comes up empty or incomplete.
    Two distinct failure modes land here: a scanned/image-only document with no
    extractable text layer at all (same root cause as CMR's handwritten containers -
    see _extract_containers_via_llm), or a real text layer whose actual label wording
    just doesn't match extract_lcl_arrival_data's patterns (a carrier/template this
    pipeline hasn't seen the exact phrasing of yet). Best-effort only; returns {} on
    any failure or on a non-multimodal-readable file, never raises."""
    if not data_bytes or not (mime or "").startswith(_LLM_READABLE_PREFIXES) or len(data_bytes) > 18 * 1024 * 1024:
        return {}
    prompt = (
        "This is an LCL (less-than-container-load) Arrival Notice or Delay / Devanning "
        "document. Read it carefully - including any handwritten or stamped text - and "
        "extract the following fields, using null for anything genuinely not present:\n"
        "- sf_number: a reference starting with \"SF\" followed by digits (e.g. SF169508)\n"
        "- container_number: an ISO-6346 container number (4 letters + 7 digits, e.g. TEMU9681744)\n"
        "- devanning_date: the specific devanning date, available date at CFS, or delayed devanning date "
        "(respond as YYYY-MM-DD). IMPORTANT: Do NOT use the vessel ETA (Estimated Time of Arrival) or document issue date as the devanning date. If there is no specific devanning, available, or delay date present, return null.\n"
        "- customs_number: the value under a label like \"Customs Number\" or "
        "\"Preceding/Previous Customs Number\" (e.g. 641761FPS-01). IMPORTANT: Do NOT return a vessel name, voyage number, ETA, \"N/A\", \"TBD\", \"available after devanning\", or any adjacent label if the customs number field is blank or missing. Return null if not present.\n"
        "- cfs_address: ONLY the short warehouse/city name on the FIRST line under a "
        "label like \"CFS Address\", \"Warehouse\", or \"Discharge CFS\" (e.g. \"VLS "
        "BELGIUM\") - do NOT include the street address, opening hours, contact "
        "emails, reference/entrepot numbers, or any other operational detail that "
        "follows it in that same block\n\n"
        "Respond with ONLY a JSON object, no markdown fences:\n"
        '{"sf_number": "...", "container_number": "...", "devanning_date": "...", '
        '"customs_number": "...", "cfs_address": "..."}'
    )
    try:
        parsed = _parse_llm_json(gemini.generate_multimodal(prompt, [(mime, data_bytes)]))
        if not parsed:
            return {}
        result = {}
        sf = re.sub(r"[^A-Z0-9]", "", str(parsed.get("sf_number") or "").upper())
        if re.match(r"^SF\d+$", sf):
            result["sf_number"] = sf
        raw_container = parsed.get("container_number")
        containers = _merge_containers([raw_container]) if raw_container else []
        if containers:
            result["container_number"] = containers[0]
        devanning = _normalize_date_str(str(parsed.get("devanning_date") or ""))
        if devanning:
            result["devanning_date"] = devanning
        if parsed.get("customs_number"):
            c_val = str(parsed["customs_number"]).strip()
            if _is_valid_customs_number(c_val):
                result["customs_number"] = c_val
        raw_cfs = str(parsed.get("cfs_address") or "").strip()
        if raw_cfs:
            # Sanity-check, not blind trust: despite the prompt's explicit instruction
            # above, Gemini has been observed grabbing the WHOLE multi-line warehouse
            # block (address + opening hours + reception email + entrepot number, HTML-
            # export "<br>" artifacts and all) instead of just the first line - the
            # same failure mode extract_lcl_arrival_data's regex fix (see its own
            # comment) already guards against on the deterministic side. A real
            # warehouse/city name is short and plain; an "@" (email), a stray "<"
            # (leftover "<br>"), or a multi-digit run beyond a couple of characters
            # (reference/entrepot/phone numbers) all indicate the answer bled past the
            # first line - truncate at the first such marker rather than trusting it
            # whole, so a bad LLM answer can't override an already-correct regex value.
            cfs_candidate = re.split(r"[<@]|\d{3,}", raw_cfs)[0].strip(" \t-:")
            if cfs_candidate and len(cfs_candidate) <= 60:
                result["cfs_address"] = cfs_candidate
        return result
    except Exception as e:
        _log_warn(f"LCL field LLM read failed for '{filename}': {e}")
        return {}


_LCL_STATUSUPDATE_RE = re.compile(r"status\s*update|\bopzetten\b|\blossen\b", re.IGNORECASE)
# "A release has been transferred" (securecontainerrelease.com-style boilerplate) is
# just a status notification ABOUT a release, not the release/delivery document itself.
_LCL_TRANSFER_RE = re.compile(r"release\s+has\s+been\s+transferred|securecontainerrelease", re.IGNORECASE)
_LCL_ARRIVAL_NOTICE_RE = re.compile(r"arrival\s*notice|\bnoa\b|notice\s*of\s*arrival", re.IGNORECASE)
_LCL_DELIVERY_ORDER_RE = re.compile(r"deliver(?:y)?\s*order|release\s*order|pin\s*sheet", re.IGNORECASE)
# "do" lowercase is too common an English filler word to trust (e.g. "please do the
# needful") - but an all-caps "DO" token is the standard reference-code convention for
# Delivery Order (e.g. subject "DO for MSCU1234567" or a filename "DO_MSCU1234567.pdf"),
# so that stricter, case-SENSITIVE form is trusted in both subject and filename.
_LCL_DO_ABBREV_RE = re.compile(r"\bDO\b")
# A carrier/terminal release-code style body line (e.g. "IMCR0101 - M;HLCULI3260631361;
# MD;ECT ;001") IS the release message itself, even with no "delivery order" wording.
_LCL_IMCR_RELEASE_LINE_RE = re.compile(r"\bimcr\d+\b", re.IGNORECASE)


def classify_lcl_arrival_email(email_obj):
    """Classify lcl-arrivals---release email into 4 distinct types based on subject,
    attachment names & body snippet:
    1. shipment_not_released (Subject contains 'shipment not released')
    2. delay_or_devanning (Subject contains 'Delay' or 'Devanning')
    3. arrival_notice - ONLY when the mail clearly is/carries an actual arrival notice
       document ("arrival notice", "NOA", "notice of arrival" wording) - never just
       because a container/booking/BL number or an ETA is mentioned in passing, or the
       mail is generically "about" a shipment.
    4. delivery_order - when the mail clearly is/carries the actual document
       ("delivery order", "release order", "pin sheet" wording, or an all-caps "DO"
       reference-code token in the subject/filename - lowercase "do" is never trusted,
       too common an English filler word), a carrier/terminal-issued release message
       listing container(s) + pickup/return details (even titled "RELEASE ORDER" - both
       map here), or a carrier release-code style body line (IMCR-format, e.g.
       "IMCR0101 - M;HLCULI3260631361;MD;ECT ;001").

    Two hard overrides run BEFORE arrival_notice/delivery_order matching, so they win
    even when the mail also carries arrival/release wording or a related attachment:
    - A transport status update (subject/body says "STATUSUPDATE", or the Dutch
      loading/unloading notices "opzetten"/"lossen") is always forced away from
      arrival_notice/delivery_order.
    - An automated "a release has been transferred" notification (e.g.
      securecontainerrelease.com boilerplate) is a status update ABOUT a release, not
      the release document itself, so it's kept out of delivery_order too.
    Both fall through to "unknown" (reviewed manually) rather than being misassigned.
    """
    subject = (email_obj.get("subject") or "").strip()
    snippet = (email_obj.get("snippet") or "").strip()
    attachment_names = email_obj.get("attachmentNames") or []
    subject_and_snippet = f"{subject} {snippet}"
    full_text = f"{subject_and_snippet} " + " ".join(attachment_names)

    if re.search(r"shipment\s+not\s+released", subject, re.IGNORECASE):
        return "shipment_not_released"
    if re.search(r"delay|devanning", subject, re.IGNORECASE):
        return "delay_or_devanning"

    if _LCL_STATUSUPDATE_RE.search(full_text) or _LCL_TRANSFER_RE.search(full_text):
        return "unknown"

    # Check for document types in attachment names first (a real attached document's
    # own filename is the most reliable signal), then subject/snippet.
    for att_name in attachment_names:
        if _LCL_ARRIVAL_NOTICE_RE.search(att_name):
            return "arrival_notice"
        if _LCL_DELIVERY_ORDER_RE.search(att_name) or _LCL_DO_ABBREV_RE.search(att_name):
            return "delivery_order"

    if _LCL_ARRIVAL_NOTICE_RE.search(subject_and_snippet):
        return "arrival_notice"
    if _LCL_DELIVERY_ORDER_RE.search(subject_and_snippet) or _LCL_DO_ABBREV_RE.search(subject):
        return "delivery_order"
    if _LCL_IMCR_RELEASE_LINE_RE.search(snippet):
        return "delivery_order"

    return "unknown"


def classify_email_meta(email, force=False, label=None):
    """READ-SAFE classify from list-view metadata only (never opens the message).

    ``email`` is a dict as produced by the scrape/fetch (id, subject, from, snippet,
    attachmentNames). ``label`` (e.g. "a-cmr") skips the general multi-type LLM/keyword
    guess entirely - every mail in that label carries a CMR document as its main
    attachment, so the first attachment (or whichever filename-hints as Cmr) is
    deterministically "Cmr" and every other real attachment is "Other", matching the
    same rule classify_documents_cmr applies to the deep/attachment-content path.
    Returns a cached record dict.
    """
    message_id = email.get("id")
    if not message_id:
        raise ValueError("email dict must include an 'id'.")

    if not force:
        cache = _load_cache()
        if message_id in cache and cache[message_id].get("source") == "metadata":
            cached = dict(cache[message_id])
            cached["cached"] = True
            return cached

    subject = email.get("subject", "") or ""
    snippet = email.get("snippet", "") or ""
    sender = email.get("from", "") or email.get("sender", "") or ""
    attachment_names = email.get("attachmentNames") or []
    # The list-row scrape sets this whenever a real attachment chip element exists on the
    # row, even if it couldn't read that chip's filename into attachment_names (a scraping
    # miss, not "no attachment") - see scripts/open_gmail.py's has_attachment computation.
    # Without this, a mail with a real but unnamed attachment fell all the way through to
    # "No DOC" below, and the downstream live-refresh escalation (tracking_api.py) never
    # opens it for a real content read because "No DOC" isn't a CONTAINER_BEARING_TYPE.
    has_attachment = bool(email.get("hasAttachment"))

    if label == "a-cmr":
        cmr_index = next((i for i, n in enumerate(attachment_names) if filename_hint(n) == "Cmr"), None)
        if cmr_index is None and attachment_names:
            cmr_index = 0
        attachments = [
            {"filename": n, "type": ("Cmr" if i == cmr_index else "Other")}
            for i, n in enumerate(attachment_names)
        ]
        if attachment_names:
            email_type, method, confidence = "Cmr", "cmr-label-override", 1.0
            reason = "Classified by a-cmr label rule."
        elif has_attachment:
            email_type, method, confidence = "Cmr", "cmr-label-override", 0.6
            reason = "Has a real attachment (filename unreadable from list view) - assumed Cmr per a-cmr label rule."
        else:
            email_type, method, confidence = "No DOC", "fallback", 0.0
            reason = "No document present."
        containers = find_container_numbers(f"{subject} {snippet} " + " ".join(attachment_names))
    elif label == _LCL_ARRIVALS_LABEL:
        # This label's document-type universe is ONLY "Arrival notice"/"Delivery
        # order"/"Other" - not the full CMR-oriented DOCUMENT_TYPES list the generic
        # branch below would otherwise guess from (which is exactly how a "B/L:" token
        # in a Delivery order subject got misclassified as "Final master bill of
        # lading", and "PLS SEND PACKINGLIST" as "Packing List"). Reuses
        # classify_lcl_arrival_email's already-tuned regex rules (the same ones that
        # route these mails to their processing handler) instead of a second, separate
        # guesser - ``email`` already has the exact subject/snippet/attachmentNames
        # shape that function expects.
        mail_type = classify_lcl_arrival_email(email)
        lcl_type = _LCL_MAIL_TYPE_TO_DOC_TYPE.get(mail_type, "Other")
        attachments = [
            {"filename": n, "type": (lcl_type if i == 0 else "Other")}
            for i, n in enumerate(attachment_names)
        ]
        if attachment_names:
            email_type, method, confidence = lcl_type, "lcl-label-override", (1.0 if lcl_type != "Other" else 0.3)
            reason = f"Classified by lcl-arrivals label rule (mail_type={mail_type})."
        elif has_attachment:
            email_type, method, confidence = lcl_type, "lcl-label-override", (0.5 if lcl_type != "Other" else 0.2)
            reason = "Has a real attachment (filename unreadable from list view) - assumed by lcl-arrivals label rule."
        else:
            email_type, method, confidence = "No DOC", "fallback", 0.0
            reason = "No document present."
        containers = find_container_numbers(f"{subject} {snippet} " + " ".join(attachment_names))
    else:
        # Keyword hint first (free, instant) - a hint and the fallback if the LLM is down.
        hint = None
        for candidate in list(attachment_names) + [subject]:
            hint = filename_hint(candidate)
            if hint:
                break

        gemini = GeminiClient()
        prompt = _build_meta_prompt(subject, snippet, sender, attachment_names, has_attachment)

        email_type, confidence, reason, method, containers = None, 0.0, "", "llm", []
        try:
            parsed = _parse_llm_json(gemini.generate_multimodal(prompt, []))
            if parsed:
                email_type = _coerce_type(parsed.get("type"))
                confidence = _safe_float(parsed.get("confidence"))
                reason = str(parsed.get("reason", ""))[:300]
                containers = parsed.get("containers") or []
        except Exception as e:
            _log_warn(f"Metadata classification LLM call failed for '{subject[:40]}': {e}")

        if not email_type:
            if hint:
                email_type, method, confidence = hint, "keyword", 0.4
                reason = reason or "Matched by filename/subject keyword."
            elif attachment_names:
                email_type, method, confidence = "Other", "fallback", 0.0
                reason = reason or "Has a document but its type is unclear from metadata."
            elif has_attachment:
                email_type, method, confidence = "Other", "fallback", 0.0
                reason = reason or "Has a real attachment but its filename couldn't be read from the list view."
            else:
                # No attachment and nothing recognisable -> don't invent a type.
                email_type, method, confidence = "No DOC", "fallback", 0.0
                reason = reason or "No document present."

        attachments = [
            {"filename": n, "type": filename_hint(n) or "Other"} for n in attachment_names
        ]
        containers = _merge_containers(containers, find_container_numbers(f"{subject} {snippet} " + " ".join(attachment_names)))

    record = {
        "message_id": message_id,
        "email_type": email_type,
        "doc_types": _distinct_types([email_type] + [a["type"] for a in attachments]),
        "type_counts": _type_counts([a["type"] for a in attachments]),
        "containers": containers,
        "subject": subject,
        "sender": sender,
        "confidence": round(confidence, 2),
        "method": method,
        "reason": reason,
        "attachments": attachments,
        "source": "metadata",
        "classified_at": datetime.now(timezone.utc).isoformat(),
    }

    with _CACHE_LOCK:
        cache = _load_cache()
        cache[message_id] = record
        _save_cache(cache)

    out = dict(record)
    out["cached"] = False
    return out


def suppress_redundant_mrn(doc_types_list, primary_type=None):
    """A body-only "MRN" (an NCTS release notification EMAIL, with no attachment of
    its own - see the classifier guardrail) is redundant once a more specific
    customs-form type is ALSO known for the same message (T1 DOC / Custome Import Doc
    - normally the actual attached document, read off its printed Declaration type
    field). Showing both is confusing, not extra information, so drop "MRN" in that
    case. Shared by every doc_types-list consumer (get_cached_type_map,
    classify_new_documents_cmr/_lcl, classify_all_documents_cmr/_lcl) so this rule
    lives in one place instead of being reimplemented slightly differently in each."""
    if "MRN" in doc_types_list and any(t in doc_types_list for t in ("T1 DOC", "Custome Import Doc")):
        filtered = [t for t in doc_types_list if t != "MRN"]
        if primary_type == "MRN" and filtered:
            primary_type = filtered[0]
        return filtered, primary_type
    return doc_types_list, primary_type


def get_cached_type_map(message_ids):
    """Return ``(types, types_all, containers, type_counts)`` for whichever of
    ``message_ids`` already have a classification cached. ``types`` is
    ``{id: primary_type}``; ``types_all`` is ``{id: [all distinct document types]}``;
    ``containers`` is ``{id: [container nos]}``; ``type_counts`` is
    ``{id: {type: instance_count}}`` (only populated once a deep/meta pass actually knows
    the per-document breakdown). Loads the cache once - used by the live auto-classify
    sweep to cheaply tell which mails are new (uncached)."""
    cache = _load_cache()
    types, types_all, containers, type_counts = {}, {}, {}, {}
    for mid in message_ids or []:
        rec = cache.get(mid)
        if rec and rec.get("email_type"):
            doc_types_list, primary = suppress_redundant_mrn(
                rec.get("doc_types") or [rec["email_type"]], rec["email_type"]
            )
            types_all[mid] = doc_types_list
            types[mid] = primary
            if rec.get("containers"):
                containers[mid] = rec["containers"]
            if rec.get("type_counts"):
                type_counts[mid] = rec["type_counts"]
    return types, types_all, containers, type_counts


def _fetch_subject(email_service, message_id):
    """Best-effort subject lookup via the read-only Gmail API (metadata only)."""
    try:
        service = email_service._get_gmail_service(["https://www.googleapis.com/auth/gmail.readonly"])
        if not service:
            return ""
        msg = service.users().messages().get(
            userId="me", id=message_id, format="metadata", metadataHeaders=["Subject"]
        ).execute()
        for h in msg.get("payload", {}).get("headers", []):
            if h.get("name", "").lower() == "subject":
                return h.get("value", "")
    except Exception:
        pass
    return ""
